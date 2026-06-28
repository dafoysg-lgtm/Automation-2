"""
Automasi Penomoran dan Rename Invoice PDF-Excel
Versi: MVP V2.4 — Hybrid Pairing
- Primary  : kode CP unik (CP\d+[A-Z]\d+) — aman, tidak bisa salah pasang
- Fallback : nomor urut di depan nama file — untuk file tanpa kode CP
- Border-safe Excel: binary copy dulu, baru edit 2 sel
- Tanya kode perusahaan saat run jika config kosong
"""

import re
import sys
import shutil
import logging
import configparser
from datetime import datetime
from pathlib import Path

try:
    import xlrd
    import xlwt
    from xlutils.copy import copy as xl_copy
    XLS_AVAILABLE = True
    XLRD_V1 = int(xlrd.__version__.split(".")[0]) < 2
except ImportError:
    XLS_AVAILABLE = False
    XLRD_V1 = False

try:
    import openpyxl
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


BASE_DIR    = Path(__file__).parent
INPUT_DIR   = BASE_DIR / "Input"
OUTPUT_DIR  = BASE_DIR / "Selesai"
LOG_DIR     = BASE_DIR / "Log"
CONFIG_FILE = BASE_DIR / "config.ini"

for d in (INPUT_DIR, OUTPUT_DIR, LOG_DIR):
    d.mkdir(exist_ok=True)

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
log_file  = LOG_DIR / f"log_{timestamp}.txt"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

BULAN_ROMAWI = {1:"I",2:"II",3:"III",4:"IV",5:"V",6:"VI",
                7:"VII",8:"VIII",9:"IX",10:"X",11:"XI",12:"XII"}
BULAN_ID     = {1:"Januari",2:"Februari",3:"Maret",4:"April",
                5:"Mei",6:"Juni",7:"Juli",8:"Agustus",
                9:"September",10:"Oktober",11:"November",12:"Desember"}

POLA_CP = r"CP\d+[A-Z]\d+"


# ─── CONFIG ───────────────────────────────────────────────────────────────────

def baca_config():
    cfg = configparser.ConfigParser()
    if CONFIG_FILE.exists():
        cfg.read(CONFIG_FILE, encoding="utf-8")
    return {
        "kode_perusahaan": cfg.get("pengaturan", "kode_perusahaan", fallback="").strip(),
        "sel_nomor_surat": cfg.get("pengaturan", "sel_nomor_surat", fallback="F11").strip(),
        "sel_tanggal":     cfg.get("pengaturan", "sel_tanggal",     fallback="F12").strip(),
    }


# ─── SCAN & HYBRID PAIRING ────────────────────────────────────────────────────

def buat_kunci(nama_file: str):
    """
    Buat kunci pairing untuk satu file.
    Priority 1: kode CP unik  → ('cp', 'CP16E00063')   ← AMAN, tidak bisa salah pasang
    Priority 2: nomor urut    → ('urut', '11')          ← fallback jika tidak ada CP
    Priority 3: nama lengkap  → ('nama', '...')         ← last resort
    """
    m_cp = re.search(POLA_CP, nama_file)
    if m_cp:
        return ("cp", m_cp.group(0))

    m_urut = re.match(r"^(\d+)\.", nama_file)
    if m_urut:
        return ("urut", m_urut.group(1))

    return ("nama", Path(nama_file).stem)


def scan_file_input():
    pasangan = {}
    konflik  = []

    for f in INPUT_DIR.iterdir():
        if not f.is_file():
            continue
        ext = f.suffix.lower()
        if ext not in (".pdf", ".xls", ".xlsx"):
            continue

        kunci = buat_kunci(f.name)

        if kunci not in pasangan:
            pasangan[kunci] = {"pdf": None, "xls": None, "mode": kunci[0]}
        else:
            slot = "pdf" if ext == ".pdf" else "xls"
            if pasangan[kunci][slot] is not None:
                konflik.append((kunci, pasangan[kunci][slot], f))
                log.error(
                    f"  ✗ KONFLIK kunci={kunci}: "
                    f"'{pasangan[kunci][slot].name}' vs '{f.name}' — keduanya dilewati"
                )
                pasangan[kunci][slot] = None
                continue

        if ext == ".pdf":
            pasangan[kunci]["pdf"] = f
        else:
            pasangan[kunci]["xls"] = f

    return pasangan, konflik


# ─── RENAME ───────────────────────────────────────────────────────────────────

def nama_file_baru(nomor_invoice: int, nama_asli: str) -> str:
    ext  = Path(nama_asli).suffix
    stem = Path(nama_asli).stem
    stem = re.sub(r"^\d+\.\s*", "", stem).strip()
    stem = re.sub(r"^Inv\s*\.+\s*", "", stem, flags=re.IGNORECASE).strip()
    return f"Inv {nomor_invoice} {stem}{ext}"


# ─── TULIS EXCEL — BORDER-SAFE ────────────────────────────────────────────────

def parse_sel(sel: str):
    m = re.match(r"([A-Za-z]+)(\d+)", sel.strip())
    if not m:
        raise ValueError(f"Format sel tidak valid: {sel}")
    col_str = m.group(1).upper()
    col = sum((ord(c) - 64) * (26 ** i)
              for i, c in enumerate(reversed(col_str))) - 1
    return int(m.group(2)) - 1, col


def tulis_excel_safe(path_src: Path, sel_nomor: str, sel_tanggal: str,
                     nomor_surat: str, tanggal: str, nama_output: str) -> Path:
    """
    1. Binary copy → format/border/merge cell 100% preserved
    2. Buka copy → edit hanya 2 sel target → simpan
    """
    ext = path_src.suffix.lower()
    dst = OUTPUT_DIR / nama_output
    shutil.copy2(str(path_src), str(dst))

    if ext == ".xlsx":
        if not XLSX_AVAILABLE:
            raise RuntimeError("Jalankan: pip install openpyxl")
        wb = openpyxl.load_workbook(str(dst))
        ws = wb.active
        ws[sel_nomor] = nomor_surat
        ws[sel_tanggal] = tanggal
        wb.save(str(dst))

    else:
        if not XLS_AVAILABLE:
            raise RuntimeError("Jalankan: pip install xlrd==1.2.0 xlwt xlutils")
        if XLRD_V1:
            rb = xlrd.open_workbook(str(dst), formatting_info=True)
            wb = xl_copy(rb)
            ws = wb.get_sheet(0)
            r_n, c_n = parse_sel(sel_nomor)
            r_t, c_t = parse_sel(sel_tanggal)
            ws.write(r_n, c_n, nomor_surat)
            ws.write(r_t, c_t, tanggal)
            wb.save(str(dst))
        else:
            try:
                wb = openpyxl.load_workbook(str(dst))
                ws = wb.active
                ws[sel_nomor] = nomor_surat
                ws[sel_tanggal] = tanggal
                wb.save(str(dst))
            except Exception:
                rb = xlrd.open_workbook(str(dst))
                wb_w = xl_copy(rb)
                ws_w = wb_w.get_sheet(0)
                r_n, c_n = parse_sel(sel_nomor)
                r_t, c_t = parse_sel(sel_tanggal)
                ws_w.write(r_n, c_n, nomor_surat)
                ws_w.write(r_t, c_t, tanggal)
                wb_w.save(str(dst))
                log.warning(
                    "  ⚠  xlrd 2.x terdeteksi — border mungkin tidak terjaga. "
                    "Fix: pip install xlrd==1.2.0"
                )
    return dst


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    log.info("═"*57)
    log.info("  AUTOMASI PENOMORAN INVOICE  — MVP V2.4")
    log.info("═"*57)

    cfg = baca_config()
    log.info(f"  Sel nomor surat : {cfg['sel_nomor_surat']}")
    log.info(f"  Sel tanggal     : {cfg['sel_tanggal']}")
    log.info(f"  Mode pairing    : Hybrid (CP kode → nomor urut → nama file)")

    log.info("\n[1/4] Scanning folder Input …")
    pasangan, konflik = scan_file_input()

    lengkap  = [(k, v) for k, v in pasangan.items() if v["pdf"] and v["xls"]]
    pdf_saja = [(k, v) for k, v in pasangan.items() if v["pdf"] and not v["xls"]]
    xls_saja = [(k, v) for k, v in pasangan.items() if v["xls"] and not v["pdf"]]

    def sort_key(item):
        k, v = item
        if k[0] == "cp":
            return (0, k[1])
        elif k[0] == "urut":
            return (1, k[1].zfill(6) if k[1].isdigit() else k[1])
        return (2, k[1])

    lengkap_sorted = sorted(lengkap, key=sort_key)

    cp_count   = sum(1 for k, v in lengkap if k[0] == "cp")
    urut_count = sum(1 for k, v in lengkap if k[0] == "urut")

    log.info(f"\n  Pasangan lengkap          : {len(lengkap)}")
    log.info(f"    └ via kode CP (aman)    : {cp_count}")
    log.info(f"    └ via nomor urut (fbk)  : {urut_count}")
    log.info(f"  PDF tanpa pasangan        : {len(pdf_saja)}")
    log.info(f"  Excel tanpa pasangan      : {len(xls_saja)}")
    if konflik:
        log.info(f"  Konflik (dilewati)        : {len(konflik)}")

    if not lengkap and not pdf_saja and not xls_saja:
        log.error("  ✗ Tidak ada file yang bisa diproses.")
        input("\n  Tekan Enter untuk keluar …")
        return

    log.info("\n[2/4] Meminta input dari user …")
    print("\n" + "═"*57)
    print("   AUTOMASI PENOMORAN INVOICE — V2.4")
    print("═"*57)

    kode_prs = cfg["kode_perusahaan"]
    if not kode_prs:
        print("\n  Pilih kode perusahaan:")
        print("    1. COMPANY_A")
        print("    2. COMPANY_B")
        print("    3. Ketik manual")
        while True:
            pilih = input("  Pilihan (1/2/3) : ").strip()
            if pilih == "1":   kode_prs = "COMPANY_A"; break
            elif pilih == "2": kode_prs = "COMPANY_B"; break
            elif pilih == "3":
                kode_prs = input("  Kode : ").strip().upper()
                if kode_prs: break
                print("  ✗ Tidak boleh kosong.")
            else:
                print("  ✗ Pilih 1, 2, atau 3.")
    else:
        print(f"\n  Kode perusahaan : {kode_prs}  (dari config.ini)")

    while True:
        try:
            nomor_awal = int(input("\n  Masukkan NOMOR AWAL invoice  (contoh: 438) : ").strip())
            break
        except ValueError:
            print("  ✗ Harus berupa angka.")

    while True:
        raw = input("  Masukkan TANGGAL INVOICE  (contoh: 24/06/2026) : ").strip()
        try:
            dt = datetime.strptime(raw, "%d/%m/%Y")
            tanggal_excel = f"{dt.day} {BULAN_ID[dt.month]} {dt.year}"
            bulan_romawi  = BULAN_ROMAWI[dt.month]
            tahun         = str(dt.year)
            break
        except ValueError:
            print("  ✗ Format salah. Gunakan DD/MM/YYYY.")

    log.info(f"  Kode perusahaan: {kode_prs}")
    log.info(f"  Nomor awal     : {nomor_awal}")
    log.info(f"  Tanggal Excel  : {tanggal_excel}")

    print(f"\n  ✓ Kode          : {kode_prs}")
    print(f"  ✓ Tanggal Excel : {tanggal_excel}")
    print(f"  ✓ Format nomor  : {nomor_awal}/{kode_prs}/{bulan_romawi}/{tahun}")

    if lengkap_sorted:
        contoh = lengkap_sorted[0][1]
        print(f"\n  Contoh PDF baru  : {nama_file_baru(nomor_awal, contoh['pdf'].name)}")
        print(f"  Contoh Excel baru: {nama_file_baru(nomor_awal, contoh['xls'].name)}")

    jml = len(lengkap_sorted)
    print(f"\n  Total pasangan  : {jml} | Nomor: {nomor_awal} s/d {nomor_awal + jml - 1}")
    if pdf_saja:
        print(f"  PDF tanpa pasangan (salin apa adanya) : {len(pdf_saja)}")
    if xls_saja:
        print(f"  Excel tanpa pasangan (salin apa adanya): {len(xls_saja)}")
    if konflik:
        print(f"  ⚠ File konflik (dilewati): {len(konflik)} — cek Log untuk detail")

    konfirm = input("\n  Lanjutkan? (y/n) : ").strip().lower()
    if konfirm != "y":
        print("  Dibatalkan.")
        return

    log.info("\n[3/4] Memproses file …\n")
    hasil_log = []
    nomor_sekarang = nomor_awal

    for kunci, berkas in lengkap_sorted:
        path_pdf = berkas["pdf"]
        path_xls = berkas["xls"]
        mode     = berkas["mode"]

        nomor_surat  = f"{nomor_sekarang}/{kode_prs}/{bulan_romawi}/{tahun}"
        nama_xls_out = nama_file_baru(nomor_sekarang, path_xls.name)
        nama_pdf_out = nama_file_baru(nomor_sekarang, path_pdf.name)
        mode_label   = f"CP:{kunci[1]}" if mode == "cp" else f"urut:{kunci[1]}"

        log.info(f"  [{nomor_sekarang}] {mode_label}")
        log.info(f"        PDF  : {path_pdf.name}")
        log.info(f"        XLS  : {path_xls.name}")
        log.info(f"        → Nomor surat : {nomor_surat}")
        log.info(f"        → Tanggal     : {tanggal_excel}")
        log.info(f"        → PDF output  : {nama_pdf_out}")
        log.info(f"        → XLS output  : {nama_xls_out}")

        try:
            tulis_excel_safe(
                path_xls, cfg["sel_nomor_surat"], cfg["sel_tanggal"],
                nomor_surat, tanggal_excel, nama_xls_out
            )
            log.info(f"        ✓ Excel disimpan : {nama_xls_out}")

            shutil.copy2(str(path_pdf), str(OUTPUT_DIR / nama_pdf_out))
            log.info(f"        ✓ PDF disalin    : {nama_pdf_out}")

            hasil_log.append({
                "nomor": nomor_surat, "mode": mode_label,
                "pdf_asal": path_pdf.name, "xls_asal": path_xls.name,
                "pdf_baru": nama_pdf_out, "xls_baru": nama_xls_out,
                "status": "OK",
            })
        except Exception as e:
            log.error(f"        ✗ GAGAL: {e}", exc_info=True)
            hasil_log.append({
                "nomor": nomor_surat, "mode": mode_label,
                "pdf_asal": path_pdf.name, "xls_asal": path_xls.name,
                "pdf_baru": "-", "xls_baru": "-",
                "status": f"GAGAL: {e}",
            })

        nomor_sekarang += 1
        log.info("")

    for _, berkas in pdf_saja:
        dst = OUTPUT_DIR / berkas["pdf"].name
        shutil.copy2(str(berkas["pdf"]), str(dst))
        log.warning(f"  ⚠  PDF tanpa pasangan disalin apa adanya: {berkas['pdf'].name}")

    for _, berkas in xls_saja:
        dst = OUTPUT_DIR / berkas["xls"].name
        shutil.copy2(str(berkas["xls"]), str(dst))
        log.warning(f"  ⚠  Excel tanpa pasangan disalin apa adanya: {berkas['xls'].name}")

    log.info("[4/4] Membuat log CSV …")
    csv_path = LOG_DIR / f"hasil_{timestamp}.csv"
    with open(csv_path, "w", encoding="utf-8-sig") as f:
        f.write("Nomor Surat,Mode Pairing,PDF Asal,XLS Asal,PDF Baru,XLS Baru,Status\n")
        for r in hasil_log:
            f.write(f"{r['nomor']},{r['mode']},{r['pdf_asal']},{r['xls_asal']},"
                    f"{r['pdf_baru']},{r['xls_baru']},{r['status']}\n")
    log.info(f"  Log CSV → {csv_path}")

    ok    = sum(1 for r in hasil_log if r["status"] == "OK")
    gagal = sum(1 for r in hasil_log if r["status"] != "OK")

    log.info("\n" + "═"*57)
    log.info(f"  SELESAI  |  Berhasil: {ok}  |  Gagal: {gagal}")
    log.info(f"  Hasil  → folder Selesai/")
    log.info(f"  Log    → folder Log/")
    log.info("═"*57)

    input("\n  Tekan Enter untuk keluar …")


if __name__ == "__main__":
    main()
